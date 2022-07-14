from openpyxl import load_workbook

wb = load_workbook("sample2.xlsx")

# print sheetnames

print(wb.sheetnames)

# select a sheet

ws = wb["marks"]

# read data in a cell

print(ws["A1"].value)

# read data row wise

for row in ws.values:
    print(row)

# read from a range

myrange = ws["C1:E3"]

for row in myrange:
    for cell in row:
        print(cell.value, end=" ")
    print()
#%%

#%%
