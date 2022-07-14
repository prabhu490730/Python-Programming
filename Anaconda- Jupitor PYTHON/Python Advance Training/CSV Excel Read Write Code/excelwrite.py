# pip install openpyxl
# available on my machine

import openpyxl
wb = openpyxl.Workbook()

ws = wb.active
ws['A1'] = "Hello"
ws['A2'] = "Prabhat"

wb.save("sample.xlsx")

#%%
