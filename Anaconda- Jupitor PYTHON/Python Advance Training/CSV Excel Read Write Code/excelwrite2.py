
import openpyxl
wb = openpyxl.Workbook() #Create a Workbook

ws = wb.create_sheet("mydata") # Select the default Sheet
ws['A1'] = "Hello"
ws['A2'] = "Prabhat"

ws2 = wb.create_sheet("marks") # Select the default sheet
ws2.append([23,45,645,34,34,54,65])
ws2.append([23,45,60,34,34,54,65])
ws2.append([23,45,65,34,44,54,65])

wb.save("sample2.xlsx") # Save the workbook
#%%
